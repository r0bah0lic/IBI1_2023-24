import xml.dom.minidom
from openpyxl import Workbook
dom = xml.dom.minidom.parse('go_obo.xml')
# Get all the term elements in the XML
term_elements = dom.getElementsByTagName('term')

# Create an Excel workbook and get the active sheet
workbook = Workbook()
sheet = workbook.active

# Write headers to the Excel sheet
sheet['A1'] = 'GO ID'
sheet['B1'] = 'Term Name'
sheet['C1'] = 'Definition'
sheet['D1'] = 'Number of Child Nodes'

# Initialize row index for writing data to the sheet
row_index = 2

# Count the number of child nodes
def count_childnodes(term_id):
    count = 0
    term_ids = [term_id]
    while term_ids:
        parent_term_id = term_ids.pop()
        for term_element in term_elements:
            is_a_elements = term_element.getElementsByTagName('is_a')
            for is_a_element in is_a_elements:
                if is_a_element.firstChild.data == parent_term_id:
                    term_ids.append(term_element.getElementsByTagName('id')[0].firstChild.data)
                    count +=1
    return count

for term_element in term_elements:
    # Get the <defstr> element within the current <term> element
    defstr_element = term_element.getElementsByTagName('defstr')[0]
    # Check if the text in <defstr> contains 'autophagosome'
    if 'autophagosome' in defstr_element.firstChild.data:
        # Get the <id>, <name> elements
        id_element = term_element.getElementsByTagName('id')[0]
        name_element = term_element.getElementsByTagName('name')[0]

        # Retrieve the GO ID, term name, and definition string
        GO_id = id_element.firstChild.data
        term_name = name_element.firstChild.data
        definition = defstr_element.firstChild.data
        count = count_childnodes(GO_id)

        # Write data to the Excel sheet
        sheet.cell(row=row_index, column=1).value = GO_id
        sheet.cell(row=row_index, column=2).value = term_name
        sheet.cell(row=row_index, column=3).value = definition
        sheet.cell(row=row_index, column=4).value = count

        row_index += 1

# Save the Excel workbook
workbook.save('beginning of the autophagosome.xlsx')
